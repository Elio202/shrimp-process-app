import sqlite3
from datetime import datetime, date, timezone, timedelta
from io import BytesIO
from pathlib import Path

_GT = timezone(timedelta(hours=-6))  # Guatemala = UTC-6, sin horario de verano


def now_gt():
    return datetime.now(_GT).strftime("%Y-%m-%d %H:%M:%S")

import pandas as pd
import streamlit as st

DB_PATH = Path(__file__).resolve().parent / "data" / "procesos.db"


def try_parse_fecha(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


DB_PATH.parent.mkdir(exist_ok=True)

AREAS = [
    "CLASIFICADO ENTERO", "CLASIFICADO COLA", "DESCABEZADO", "DESCONGELADO",
    "EMPAQUE", "MASTERIZADO", "PELADO", "TRATAMIENTO", "TUNEL IQF",
    "REEMPAQUE", "REETIQUETADO",
]

NO_APLICA_AREAS = {"CLASIFICADO COLA", "PELADO", "DESCABEZADO", "TRATAMIENTO", "DESCONGELADO"}
NO_PRESENTACION_AREAS = {"CLASIFICADO COLA", "PELADO", "DESCABEZADO", "TRATAMIENTO"}

_PISCINAS_EXTRANJERO = {"Q001", "Q002", "Q003", "Q004", "Q005"}

PRODUCTOS = [
    "GRANEL", "DESCOMPUESTO", "BASURA", "DESCABEZADO", "ENTERO", "ENTERO-SIFON",
    "ENTERO-IQF", "ENTERO-R", "ENTERO 100% MUDADO", "COLA", "COLA-SIFON", "COLA-IQF",
    "COLA-R", "P&D T-ON", "P&D T-OFF", "P&D T-OFF CC", "P&D T-OFF R", "PPV T-ON",
    "PPV T-OFF", "PPV T-OFF CC", "PUD T-ON", "PUD T-OFF", "PUD T-OFF R", "PUD T-OFF CC",
    "BF T-ON", "BF T-OFF", "BF T-OFF R", "BF T-ON ALINEADO", "3/4 T-ON", "3/4 T-OFF",
    "3/4 T-ON CC", "BUTTER ROUND PyD T-ON", "PINCHOS T-ON", "PINCHOS T-OFF", "PINCHOS BF",
    "PINCHOS BF ENTERO", "ENTERO BF", "COCINADO ENTERO", "COCINADO COLA", "COCINADO EZ-PEEL",
    "COCINADO P&D T-ON", "COCINADO P&D T-OFF", "COCINADO P&D T-OFF CC", "COCINADO PPV T-ON",
    "COCINADO PPV T-OFF", "COCINADO PPV T-OFF CC", "COCINADO PUD T-ON", "COCINADO PUD T-OFF",
    "COCINADO PUD T-OFF CC", "EMPANIZADO P&D T-ON", "EMPANIZADO P&D T-OFF", "EMPANIZADO PPV T-ON",
    "EMPANIZADO PPV T-OFF", "EMPANIZADO PUD T-ON", "EMPANIZADO PUD T-OFF", "EMPANIZADO BF T-ON",
    "PPV T-OFF R", "SP-CASCARA", "SP-CASCARA Y CABEZA",
]

CLIENTES = [
    "OROPSA", "OFICINAS CENTRALES", "PRODUCTO PARA PELADO", "YENS", "LEQUALITY",
    "I OCEAN", "PESCADORES", "GOLD LAKE", "GOLD LAKE BELLA", "UNION MARINE",
    "RED CHAMBER COMPANY", "PESCANOVA USA", "INDUPECASA", "WALMART", "SUMINISTROS",
    "UNISUPER", "VENTA LOCAL", "SIN CLIENTE", "LAI LAI", "RINOLI S.A",
    "(pendiente de definir)",
    "IMPORTADORA Y EXPORTADORA DE MARISCOS DE CENTRO AMERICA Y EL CARIBE S.A DE CV",
    "EL HUEVO FRITO S.A.", "OCEANA",
]

TALLAS = [
    "6 OZ", "4 OZ", "8 OZ", "L1", "L2", "L1 11/20", "U/10", "U/15", "10/20", "U/8",
    "21/25", "20/30", "31/35", "30/40", "40/50", "50/60", "60/70", "70/80", "80/100",
    "90/100", "100/120", "120/150", "150/200", "200/300", "300/400", "33/37", "C1 30/55",
    "ROTA", "56/100", "MIX", "C1", "C2", "16/20", "21/30", "26/30", "31/40", "36/40",
    "41/50", "51/60", "61/70", "80/120", "71/90", "91/110", "91/120", "110/130",
    "130/150", "200/300",
    "1 GRAMOS", "2 GRAMOS", "3 GRAMOS", "4 GRAMOS", "5 GRAMOS", "6 GRAMOS",
    "7 GRAMOS", "8 GRAMOS", "9 GRAMOS", "10 GRAMOS", "11 GRAMOS", "12 GRAMOS",
    "13 GRAMOS", "14 GRAMOS", "15 GRAMOS", "16 GRAMOS", "17 GRAMOS", "18 GRAMOS",
    "19 GRAMOS", "20 GRAMOS", "21 GRAMOS", "22 GRAMOS", "23 GRAMOS", "24 GRAMOS",
    "25 GRAMOS", "26 GRAMOS", "27 GRAMOS", "28 GRAMOS", "29 GRAMOS", "30 GRAMOS",
    "31 GRAMOS", "32 GRAMOS", "33 GRAMOS", "34 GRAMOS", "35 GRAMOS", "36 GRAMOS",
    "37 GRAMOS", "38 GRAMOS", "39 GRAMOS", "40 GRAMOS", "41 GRAMOS", "42 GRAMOS",
    "43 GRAMOS", "44 GRAMOS", "45 GRAMOS", "46 GRAMOS", "47 GRAMOS", "48 GRAMOS",
    "49 GRAMOS", "50 GRAMOS", "51 GRAMOS", "52 GRAMOS", "53 GRAMOS", "54 GRAMOS",
    "55 GRAMOS", "56 GRAMOS", "57 GRAMOS", "58 GRAMOS", "59 GRAMOS", "60 GRAMOS",
    "66 GRAMOS", "SIN TALLA", "BROKEN", "SMALL BROKEN", "MEDIUM BROKEN", "BIG BROKEN",
    "PEQUEÑO (60/70 - 150/200)", "MEDIANO (30/40 - 50/60)", "GRANDE",
]

PRESENTACIONES = [
    "1/1 kg (1 kg)", "2/10 kg (20 kg)", "10/3 lb (30 lb)", "7/4 lb (28 lb)",
    "19/1.18 lb (22..42 lb)", "20/0.5 lb(10lb)", "32/4-6 oz (10lb)", "40/4 oz (10 lb)",
    "26/6 oz (10 lb)", "20/8 oz (10 lb)", "16/10 oz (10 lb)", "40/ 8 oz(20 lb)",
    "1/10 (10 kg)", "10/1-3 oz (10 lb)", "5/4 lb (20 lb)", "3/6 kg (18 kg)",
    "30/1 lb (30 lb)", "7/2 KG", "19/1.05 lb (20 lb)", "24/450 gr (10.89 kg)",
    "18/1.11 lb (20 lb)", "12/450 gr (5.4 kg)", "24/450 gr (10.8 kg)", "16/450 gr (7.2 kg)",
    "22/450 gr (9.9 kg)", "20/450 gr(9 kg)", "12/480 gr(5.76kg)", "16/480 gr (7.68 kg)",
    "24/454 gr (10.89 kg)", "22/480 gr (10.56 kg)", "12/600 gr(7.2 kg)", "14/650 gr (9.1 kg)",
    "14/0.454 kg (6.35 kg)", "22/650 gr (14.3 kg)", "12/650 gr (7.7 kg)", "22/680 gr (14.96 kg)",
    "24/600 gr(14.4 kg)", "16/650 gr (10.4 kg)", "20/650 gr (13 kg)", "14/680 gr (9.52 kg)",
    "19/523 gr (9.9 kg)", "10/ 720 gr (7.2 kg)", "5/2 kg(10kg)", "10/1 lb (10 lb)",
    "5/2 lb (10 lb)", "4/5 lb (20 lb)", "1/15 lb (15 lb)", "25/1 lb (25 lb)",
    "6/3 lb (18 lb)", "9/2 lb (18 lb)", "10/2 lb (20 lb)", "4/3 lb (12 lb)",
    "20/1 lb (20 lb)", "20/340 gr (6.8 kg)", "14/1.5 lb (21 lb)", "6/4 lb (24 lb)",
    "6/4.4 lb (26.4 lb)", "14/2 lb (28 lb)", "1/30 lb (30 lb)", "10/4 lb (40 lb)",
    "1/20 lb (20 lb)", "2/20 lb (40 lb)", "1/25 lb (25lb)", "1/40 lb (40 lb)",
    "10/5 lb (50 lb)", "1/50 lb (50 lb)", "24/1 lb (24 lb)", "10/720 gr(7.2 kg)",
    "12/1 LB (12 lb)", "19/500 gr (9.5kg)", "20/375 gr (7.5 kg)", "14/700 gr (9.8kg)",
    "20/360 gr(7.2 kg)", "1/10 lb (10 lb)", "25/360 gr (9 kg)", "10/2 kg (20 kg)",
    "20/420 gr (8.4 kg)", "10/908 gr (9.08 kg)", "6/2 kg (12 kg)", "8/2 kg (16 kg)",
    "14/3 lb (42 lb)", "4/2 Kg (8 Kg)", "6/5 lb (30 lb)", "8/5 lb (40 lb)",
    "6/1 lb (6 lb)", "6/2 lb (12 lb)", "8/4 lb (32 lb)",
]

POOLS = [
    "TM01-E01","TM01-E02","TM01-E03","TM01-E04","TM01-P01","TM01-P02","TM01-P03","TM00-P0A",
    "TM02-E01","TM02-E02","TM02-E03","TM02-E04","TM02-E05","TM02-P01","TM02-P02","TM02-P03","TM02-P04",
    "TM03-E01","TM03-E02","TM03-E03","TM03-E04","TM03-P01",
    "TM04-E01","TM04-E02","TM04-E03","TM04-E04","TM04-E05","TM04-E06","TM04-E07",
    "TM05-E01","TM05-E02","TM05-E03","TM05-E04","TM05-E05","TM05-E06","TM05-P01","TM05-P02","TM05-P03",
    "TM06-E01","TM06-E02","TM06-E03","TM06-E04","TM06-E05","TM06-E06",
    "EM00-PC1","EM00-PC2","EM00-PC3","EM00-PC4",
    "EM01-E01","EM01-E02","EM01-E03","EM01-E04",
    "EM02-E01","EM02-E02","EM02-E03","EM02-E04","EM02-E05",
    "EM03-E01","EM03-E02","EM03-E03","EM03-E04","EM03-E05","EM03-P01","EM03-P02","EM03-P03","EM03-P04",
    "EM04-E01","EM04-E02","EM04-E03","EM04-E04","EM04-E05","EM04-E06","EM04-E07","EM04-E08",
    "EM04-P01","EM04-P02","EM04-P03","EM04-P04",
    "EM05-E01","EM05-E02","EM05-E03","EM05-E04","EM05-E05","EM05-E06","EM05-P01","EM05-P02","EM05-P03",
    "EM06-E01","EM06-E02","EM06-E03","EM06-E04","EM06-E05","EM06-E06",
    "Q001","Q002","Q003","Q004","Q005",
]

CICLOS = [str(i) for i in range(1, 31)]


# ── Database ───────────────────────────────────────────────────────────────────

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_connection() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ingresos (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha       TEXT NOT NULL,
                turno       TEXT NOT NULL,
                area        TEXT NOT NULL,
                kg_recibidos REAL DEFAULT 0,
                cliente     TEXT DEFAULT '',
                no_personas INTEGER DEFAULT 0,
                observaciones TEXT DEFAULT '',
                creado_en   TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS detalles (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                ingreso_id   INTEGER NOT NULL,
                lote_codigo  TEXT DEFAULT '',
                piscina      TEXT DEFAULT '',
                ciclo        INTEGER,
                lote         TEXT DEFAULT '',
                tipo_producto TEXT DEFAULT '',
                tallas       TEXT DEFAULT '',
                presentacion TEXT DEFAULT 'No aplica',
                kg_procesados REAL DEFAULT 0
            )
        """)

        # Migrate from legacy registros table if present
        has_registros = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='registros'"
        ).fetchone()
        if has_registros:
            n_ing = conn.execute("SELECT COUNT(*) as n FROM ingresos").fetchone()["n"]
            if n_ing == 0:
                reg_cols = {r["name"] for r in conn.execute("PRAGMA table_info(registros)").fetchall()}
                for col, ctype in [
                    ("lote_codigo", "TEXT"), ("piscina", "TEXT"), ("ciclo", "INTEGER"),
                    ("kg_recibidos", "REAL"), ("kg_procesados", "REAL"), ("no_personas", "INTEGER"),
                    ("observaciones", "TEXT"), ("tipo_producto", "TEXT"), ("tallas", "TEXT"),
                    ("presentacion", "TEXT"), ("lote", "TEXT"), ("cliente", "TEXT"),
                ]:
                    if col not in reg_cols:
                        conn.execute(f"ALTER TABLE registros ADD COLUMN {col} {ctype}")
                conn.execute("""
                    INSERT INTO ingresos (id, fecha, turno, area, kg_recibidos, cliente, no_personas, observaciones, creado_en)
                    SELECT id, fecha, turno, area,
                           COALESCE(kg_recibidos, 0), COALESCE(cliente, ''),
                           COALESCE(no_personas, 0), COALESCE(observaciones, ''),
                           COALESCE(creado_en, datetime('now'))
                    FROM registros
                """)
                conn.execute("""
                    INSERT INTO detalles (ingreso_id, lote_codigo, piscina, ciclo, lote, tipo_producto, tallas, presentacion, kg_procesados)
                    SELECT id, COALESCE(lote_codigo,''), COALESCE(piscina,''), ciclo,
                           COALESCE(lote,''), COALESCE(tipo_producto,''),
                           COALESCE(tallas,''), COALESCE(presentacion,'No aplica'),
                           COALESCE(kg_procesados, 0)
                    FROM registros
                """)
                conn.execute("ALTER TABLE registros RENAME TO registros_v1")
        conn.commit()


def insertar_ingreso(fecha, turno, area, kg_recibidos, cliente, no_personas, observaciones):
    with get_connection() as conn:
        cur = conn.execute(
            """INSERT INTO ingresos (fecha, turno, area, kg_recibidos, cliente, no_personas, observaciones, creado_en)
               VALUES (?,?,?,?,?,?,?,?)""",
            (fecha, turno, area, kg_recibidos, cliente, no_personas, observaciones,
             now_gt()),
        )
        conn.commit()
        return cur.lastrowid


def insertar_detalle(ingreso_id, lote_codigo, piscina, ciclo, lote, tipo_producto, tallas, presentacion, kg_procesados):
    with get_connection() as conn:
        conn.execute(
            """INSERT INTO detalles (ingreso_id, lote_codigo, piscina, ciclo, lote, tipo_producto, tallas, presentacion, kg_procesados)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (ingreso_id, lote_codigo, piscina, ciclo, lote, tipo_producto, tallas, presentacion, kg_procesados),
        )
        conn.commit()


def get_registros_joined(area_filter="Todas"):
    with get_connection() as conn:
        sql = """
            SELECT i.id as ingreso_id, d.id as detalle_id,
                   i.fecha, i.turno, i.area, i.kg_recibidos, i.cliente,
                   i.no_personas, i.observaciones, i.creado_en,
                   d.lote_codigo, d.piscina, d.ciclo, d.lote,
                   d.tipo_producto, d.tallas, d.presentacion, d.kg_procesados
            FROM ingresos i
            LEFT JOIN detalles d ON d.ingreso_id = i.id
        """
        if area_filter and area_filter != "Todas":
            rows = conn.execute(sql + " WHERE i.area=? ORDER BY i.fecha DESC, i.id DESC, d.id", (area_filter,)).fetchall()
        else:
            rows = conn.execute(sql + " ORDER BY i.fecha DESC, i.id DESC, d.id").fetchall()
    return [dict(r) for r in rows]


def get_ingresos(area_filter="Todas"):
    with get_connection() as conn:
        if area_filter and area_filter != "Todas":
            rows = conn.execute(
                "SELECT * FROM ingresos WHERE area=? ORDER BY fecha DESC, id DESC", (area_filter,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM ingresos ORDER BY fecha DESC, id DESC").fetchall()
    return [dict(r) for r in rows]


def get_detalles_by_ingreso(ingreso_id):
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM detalles WHERE ingreso_id=? ORDER BY id", (ingreso_id,)
        ).fetchall()
    return [dict(r) for r in rows]


def actualizar_ingreso(ingreso_id, fecha, turno, area, kg_recibidos, cliente, no_personas, observaciones):
    with get_connection() as conn:
        conn.execute(
            """UPDATE ingresos SET fecha=?,turno=?,area=?,kg_recibidos=?,
               cliente=?,no_personas=?,observaciones=? WHERE id=?""",
            (fecha, turno, area, kg_recibidos, cliente, no_personas, observaciones, ingreso_id),
        )
        conn.commit()


def actualizar_detalle(detalle_id, lote_codigo, piscina, ciclo, lote, tipo_producto, tallas, presentacion, kg_procesados):
    with get_connection() as conn:
        conn.execute(
            """UPDATE detalles SET lote_codigo=?,piscina=?,ciclo=?,lote=?,
               tipo_producto=?,tallas=?,presentacion=?,kg_procesados=? WHERE id=?""",
            (lote_codigo, piscina, ciclo, lote, tipo_producto, tallas, presentacion, kg_procesados, detalle_id),
        )
        conn.commit()


def eliminar_ingreso(ingreso_id):
    with get_connection() as conn:
        conn.execute("DELETE FROM detalles WHERE ingreso_id=?", (ingreso_id,))
        conn.execute("DELETE FROM ingresos WHERE id=?", (ingreso_id,))
        conn.commit()


# ── Page setup ─────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Control de planta de camarón", layout="wide")
st.markdown("""
    <style>
    .stApp { background-color: #b8d0eb; }
    [data-testid="stImage"] > img { background: transparent !important; }
    </style>
""", unsafe_allow_html=True)

_col_title, _col_logo = st.columns([3, 1])
with _col_title:
    st.title("Planta de Proceso OROPSA")
    st.markdown("<p style='font-size:20px;'>Sistema de manejo de información en planta procesadora.</p>",
                unsafe_allow_html=True)
with _col_logo:
    st.image("LogoV2-removebg-preview.png", use_container_width=True)

init_db()

# ── Sidebar ────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.header("Áreas")
    selected_area = st.selectbox("Selecciona el área", AREAS, index=0)
    st.caption("Elige el área desde la lista para registrar la producción.")
    st.divider()
    area_filter = st.selectbox("Filtrar historial", ["Todas", *AREAS])

# ── Summary ────────────────────────────────────────────────────────────────────

registros_join = get_registros_joined(area_filter)
if registros_join:
    _df_sum = pd.DataFrame(registros_join)
    _total_kgp = _df_sum["kg_procesados"].fillna(0).sum()
    _total_ing = int(_df_sum["ingreso_id"].nunique())
else:
    _total_kgp = 0
    _total_ing = 0

st.subheader("Resumen")
_sc1, _sc2, _sc3 = st.columns(3)
_sc1.metric("Ingresos registrados", _total_ing)
_sc2.metric("Área seleccionada", selected_area)
_sc3.metric("KG procesados", f"{_total_kgp:,.0f}")

st.divider()

# ── Capture Form (maestro-detalle) ─────────────────────────────────────────────

with st.container(border=True):
    st.subheader(f"Nuevo ingreso — {selected_area}")

    if "form_ver" not in st.session_state:
        st.session_state["form_ver"] = 0
    if "num_lineas" not in st.session_state:
        st.session_state["num_lineas"] = 1

    v = st.session_state["form_ver"]
    n = st.session_state["num_lineas"]

    # ── Encabezado ──
    h1, h2, h3 = st.columns([2, 1, 1])
    with h1:
        fecha_text = st.text_input(
            "Fecha", value=date.today().strftime("%d/%m/%Y"),
            key=f"h_fecha_{v}", placeholder="dd/mm/aaaa"
        )
    with h2:
        turno = st.selectbox("Turno", ["Mañana", "Tarde", "Noche", "Otro"], key=f"h_turno_{v}")
    with h3:
        no_personas = st.number_input("No. personas", min_value=0, step=1, format="%d", key=f"h_personas_{v}")

    h4, h5 = st.columns(2)
    with h4:
        kg_recibidos = st.number_input(
            "KG recibidos (total del ingreso)", min_value=0.0, step=1.0, key=f"h_kgr_{v}"
        )
    with h5:
        if selected_area in NO_APLICA_AREAS:
            cliente = "No aplica"
            st.text_input("Cliente", value=cliente, disabled=True, key=f"h_cli_d_{v}")
        else:
            _cl_sel = st.selectbox("Cliente", ["Selecciona", *CLIENTES], key=f"h_cli_{v}")
            cliente = _cl_sel if _cl_sel != "Selecciona" else ""

    st.markdown("---")
    st.markdown("##### Detalle de producción")

    lineas_data = []
    for i in range(n):
        with st.container(border=True):
            st.caption(f"Línea {i + 1}")

            la, lb, lc_ = st.columns([2, 2, 1])
            with la:
                lote_codigo = st.text_input(
                    "Lote del día", key=f"l_cod_{v}_{i}", placeholder="Ej. G326"
                )
            with lb:
                _psc_sel = st.selectbox("Piscina", ["Selecciona", *POOLS], key=f"l_psc_{v}_{i}")
                piscina = _psc_sel if _psc_sel != "Selecciona" else ""
            with lc_:
                if piscina in _PISCINAS_EXTRANJERO:
                    st.text_input("Ciclo", value="No aplica", disabled=True, key=f"l_cic_d_{v}_{i}")
                    ciclo = None
                else:
                    _cic_sel = st.selectbox("Ciclo", ["Selecciona", *CICLOS], key=f"l_cic_{v}_{i}")
                    ciclo = int(_cic_sel) if _cic_sel not in ("Selecciona", "") else None

            _parts = [p for p in [lote_codigo, piscina, str(ciclo) if ciclo else ""] if p]
            lote_uni = "-".join(_parts)
            if lote_uni:
                st.caption(f"Lote unificado: **{lote_uni}**")

            ld, le, lf, lg = st.columns([2, 2, 2, 1.5])
            with ld:
                _tp_sel = st.selectbox("Tipo de producto", ["Selecciona", *PRODUCTOS], key=f"l_tp_{v}_{i}")
                tipo_producto = _tp_sel if _tp_sel != "Selecciona" else ""
            with le:
                _ta_sel = st.selectbox("Tallas", ["Selecciona", *TALLAS], key=f"l_ta_{v}_{i}")
                tallas = _ta_sel if _ta_sel != "Selecciona" else ""
            with lf:
                if selected_area in NO_PRESENTACION_AREAS:
                    presentacion = "No aplica"
                    st.text_input("Presentación", value=presentacion, disabled=True, key=f"l_pr_d_{v}_{i}")
                else:
                    _pr_sel = st.selectbox("Presentación", ["Selecciona", *PRESENTACIONES], key=f"l_pr_{v}_{i}")
                    presentacion = _pr_sel if _pr_sel != "Selecciona" else ""
            with lg:
                kgp = st.number_input("KG procesados", min_value=0.0, step=0.1, key=f"l_kgp_{v}_{i}")

            lineas_data.append({
                "lote_codigo": lote_codigo, "piscina": piscina, "ciclo": ciclo,
                "lote": lote_uni, "tipo_producto": tipo_producto, "tallas": tallas,
                "presentacion": presentacion or "No aplica", "kg_procesados": kgp,
            })

    # Totales reactivos
    total_kgp_cap = sum(l["kg_procesados"] for l in lineas_data)
    rend_cap = round(total_kgp_cap / kg_recibidos * 100, 2) if kg_recibidos > 0 else 0.0

    mt1, mt2 = st.columns(2)
    mt1.metric("Total KG procesados", f"{total_kgp_cap:,.2f}")
    mt2.metric("Rendimiento calculado", f"{rend_cap:.2f}%")

    observaciones = st.text_area("Observaciones", key=f"h_obs_{v}")

    btn1, btn2, btn3 = st.columns([1, 1, 5])
    with btn1:
        if st.button("＋ Línea", disabled=(n >= 12), help="Agregar línea de detalle (máx. 12)"):
            st.session_state["num_lineas"] = min(n + 1, 12)
            st.rerun()
    with btn2:
        if st.button("Nuevo", help="Limpiar el formulario"):
            st.session_state["num_lineas"] = 1
            st.session_state["form_ver"] = v + 1
            st.rerun()
    with btn3:
        if st.button("Guardar ingreso", type="primary"):
            fecha_obj = try_parse_fecha(fecha_text)
            if fecha_obj is None:
                st.error("Ingresa la fecha en formato dd/mm/aaaa.")
            elif total_kgp_cap == 0:
                st.error("Ingresa al menos una línea con KG procesados.")
            else:
                ing_id = insertar_ingreso(
                    fecha_obj.strftime("%Y-%m-%d"), turno, selected_area,
                    float(kg_recibidos), cliente, int(no_personas), observaciones,
                )
                for l in lineas_data:
                    if l["kg_procesados"] > 0 or l["lote_codigo"]:
                        insertar_detalle(
                            ing_id, l["lote_codigo"], l["piscina"], l["ciclo"],
                            l["lote"], l["tipo_producto"], l["tallas"], l["presentacion"],
                            float(l["kg_procesados"]),
                        )
                st.session_state["num_lineas"] = 1
                st.session_state["form_ver"] = v + 1
                st.toast(f"Ingreso #{ing_id} guardado — Rendimiento: {rend_cap:.2f}%", icon="✅")
                st.rerun()

st.divider()

# ── Editar / Eliminar ──────────────────────────────────────────────────────────

st.subheader("Editar o eliminar ingresos")
ingresos_list = get_ingresos(area_filter)

if ingresos_list:
    ing_by_id = {i["id"]: i for i in ingresos_list}
    sel_ing_id = st.selectbox(
        "Selecciona un ingreso",
        options=list(ing_by_id.keys()),
        format_func=lambda id_: (
            f"#{id_} — {ing_by_id[id_]['fecha']} | {ing_by_id[id_]['area']} | "
            f"KGR: {ing_by_id[id_]['kg_recibidos']:.1f} kg"
        ),
    )
    sel_ing = ing_by_id[sel_ing_id]
    detalles_sel = get_detalles_by_ingreso(sel_ing_id)

    with st.form("editar_ingreso"):
        st.markdown("**Encabezado del ingreso**")
        ec1, ec2, ec3 = st.columns([2, 1, 1])
        with ec1:
            e_fecha = st.text_input(
                "Fecha",
                value=datetime.strptime(sel_ing["fecha"], "%Y-%m-%d").strftime("%d/%m/%Y"),
                placeholder="dd/mm/aaaa",
            )
        with ec2:
            _turnos = ["Mañana", "Tarde", "Noche", "Otro"]
            e_turno = st.selectbox(
                "Turno", _turnos,
                index=_turnos.index(sel_ing["turno"]) if sel_ing["turno"] in _turnos else 0,
            )
        with ec3:
            e_no_personas = st.number_input(
                "No. personas", min_value=0, step=1, format="%d",
                value=int(sel_ing["no_personas"] or 0),
            )

        ec4, ec5 = st.columns(2)
        with ec4:
            e_area = st.selectbox(
                "Área", AREAS,
                index=AREAS.index(sel_ing["area"]) if sel_ing["area"] in AREAS else 0,
            )
            e_kg_recibidos = st.number_input(
                "KG recibidos", min_value=0.0, step=1.0, value=float(sel_ing["kg_recibidos"] or 0)
            )
        with ec5:
            if e_area in NO_APLICA_AREAS:
                e_cliente = "No aplica"
                st.text_input("Cliente", value=e_cliente, disabled=True)
            else:
                _ecl_opts = ["Selecciona", *CLIENTES]
                _ecl_idx = _ecl_opts.index(sel_ing["cliente"]) if sel_ing.get("cliente") in _ecl_opts else 0
                _ecl_sel = st.selectbox("Cliente", _ecl_opts, index=_ecl_idx)
                e_cliente = _ecl_sel if _ecl_sel != "Selecciona" else ""

        e_observaciones = st.text_area("Observaciones", value=sel_ing.get("observaciones") or "")

        # Líneas de detalle editables
        lineas_edit = []
        if detalles_sel:
            st.markdown("**Líneas de detalle**")
            for det in detalles_sel:
                with st.container(border=True):
                    st.caption(f"Línea #{det['id']}")
                    de1, de2, de3 = st.columns([2, 2, 1])
                    with de1:
                        de_lc = st.text_input(
                            "Lote del día", value=det.get("lote_codigo") or "", key=f"e_lc_{det['id']}"
                        )
                    with de2:
                        _dp_opts = ["Selecciona", *POOLS]
                        _dp_idx = _dp_opts.index(det.get("piscina")) if det.get("piscina") in _dp_opts else 0
                        _dp_sel = st.selectbox("Piscina", _dp_opts, index=_dp_idx, key=f"e_psc_{det['id']}")
                        de_psc = _dp_sel if _dp_sel != "Selecciona" else ""
                    with de3:
                        _dc_opts = ["Selecciona", *CICLOS]
                        _stored_cic = str(det.get("ciclo")) if det.get("ciclo") else ""
                        _dc_idx = _dc_opts.index(_stored_cic) if _stored_cic in _dc_opts else 0
                        _dc_sel = st.selectbox("Ciclo", _dc_opts, index=_dc_idx, key=f"e_cic_{det['id']}")
                        de_cic = int(_dc_sel) if _dc_sel not in ("Selecciona", "") else None

                    de4, de5, de6, de7 = st.columns([2, 2, 2, 1.5])
                    with de4:
                        _dt_opts = ["Selecciona", *PRODUCTOS]
                        _dt_idx = _dt_opts.index(det.get("tipo_producto")) if det.get("tipo_producto") in _dt_opts else 0
                        _dt_sel = st.selectbox("Tipo de producto", _dt_opts, index=_dt_idx, key=f"e_tp_{det['id']}")
                        de_tipo = _dt_sel if _dt_sel != "Selecciona" else ""
                    with de5:
                        _dta_opts = ["Selecciona", *TALLAS]
                        _dta_idx = _dta_opts.index(det.get("tallas")) if det.get("tallas") in _dta_opts else 0
                        _dta_sel = st.selectbox("Tallas", _dta_opts, index=_dta_idx, key=f"e_ta_{det['id']}")
                        de_talla = _dta_sel if _dta_sel != "Selecciona" else ""
                    with de6:
                        if e_area in NO_PRESENTACION_AREAS:
                            de_pres = "No aplica"
                            st.text_input("Presentación", value=de_pres, disabled=True, key=f"e_pr_d_{det['id']}")
                        else:
                            _dpr_opts = ["Selecciona", *PRESENTACIONES]
                            _dpr_idx = _dpr_opts.index(det.get("presentacion")) if det.get("presentacion") in _dpr_opts else 0
                            _dpr_sel = st.selectbox("Presentación", _dpr_opts, index=_dpr_idx, key=f"e_pr_{det['id']}")
                            de_pres = _dpr_sel if _dpr_sel != "Selecciona" else ""
                    with de7:
                        de_kgp = st.number_input(
                            "KG procesados", min_value=0.0, step=0.1,
                            value=float(det.get("kg_procesados") or 0), key=f"e_kgp_{det['id']}"
                        )

                    _de_parts = [p for p in [de_lc, de_psc, str(de_cic) if de_cic else ""] if p]
                    de_lote = "-".join(_de_parts)
                    lineas_edit.append({
                        "id": det["id"], "lote_codigo": de_lc, "piscina": de_psc,
                        "ciclo": de_cic, "lote": de_lote, "tipo_producto": de_tipo,
                        "tallas": de_talla, "presentacion": de_pres or "No aplica",
                        "kg_procesados": de_kgp,
                    })

        btn_upd = st.form_submit_button("Actualizar ingreso")
        if btn_upd:
            e_fecha_obj = try_parse_fecha(e_fecha)
            if e_fecha_obj is None:
                st.error("Ingresa la fecha en formato dd/mm/aaaa.")
            else:
                actualizar_ingreso(
                    sel_ing_id, e_fecha_obj.strftime("%Y-%m-%d"),
                    e_turno, e_area, float(e_kg_recibidos),
                    e_cliente, int(e_no_personas), e_observaciones,
                )
                for l in lineas_edit:
                    actualizar_detalle(
                        l["id"], l["lote_codigo"], l["piscina"], l["ciclo"],
                        l["lote"], l["tipo_producto"], l["tallas"], l["presentacion"],
                        float(l["kg_procesados"]),
                    )
                st.success("Ingreso actualizado correctamente")
                st.rerun()

    if st.button("Eliminar ingreso seleccionado", type="secondary"):
        eliminar_ingreso(sel_ing_id)
        st.success("Ingreso eliminado")
        st.rerun()
else:
    st.info("Aún no hay ingresos para editar o eliminar.")

st.divider()

# ── Historial ──────────────────────────────────────────────────────────────────

st.subheader("Historial")
if registros_join:
    display_df = pd.DataFrame(registros_join)
    display_df["fecha"] = pd.to_datetime(display_df["fecha"]).dt.strftime("%Y-%m-%d")

    cols_show = [
        "ingreso_id", "detalle_id", "fecha", "turno", "area", "kg_recibidos", "cliente",
        "lote_codigo", "piscina", "ciclo", "lote", "tipo_producto", "tallas", "presentacion",
        "kg_procesados", "no_personas", "observaciones", "creado_en",
    ]
    cols_show = [c for c in cols_show if c in display_df.columns]
    st.dataframe(display_df[cols_show], use_container_width=True)

    # Excel con kg_proporcional correcto (agrupado por ingreso_id)
    excel_df = display_df[cols_show].copy()
    _total_por_ingreso = excel_df.groupby("ingreso_id")["kg_procesados"].transform("sum")
    excel_df["kg_proporcional"] = (
        excel_df["kg_procesados"] / _total_por_ingreso.replace(0, float("nan")) * excel_df["kg_recibidos"]
    ).round(2)
    # Guardar como decimal (0.4929) para que Excel lo trate como número porcentaje nativo
    excel_df["rendimiento_proporcional"] = (
        excel_df["kg_proporcional"] / excel_df["kg_recibidos"].replace(0, float("nan"))
    ).round(4)

    cols_excel = [
        "ingreso_id", "detalle_id", "fecha", "turno", "area", "cliente",
        "lote_codigo", "piscina", "ciclo", "lote", "tipo_producto", "tallas", "presentacion",
        "kg_recibidos", "kg_procesados", "kg_proporcional", "rendimiento_proporcional",
        "no_personas", "observaciones", "creado_en",
    ]
    cols_excel = [c for c in cols_excel if c in excel_df.columns]

    buf = BytesIO()
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        excel_df[cols_excel].to_excel(writer, index=False, sheet_name="Registros")
        ws = writer.sheets["Registros"]
        rend_idx = cols_excel.index("rendimiento_proporcional") + 1
        for cell in ws[get_column_letter(rend_idx)][1:]:
            cell.number_format = "0.00%"
    st.download_button(
        label="Descargar Excel",
        data=buf.getvalue(),
        file_name="registros_planta.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("No hay datos para mostrar todavía.")
